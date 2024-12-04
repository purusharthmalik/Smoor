import os
from xlrd import open_workbook
from openpyxl import Workbook

def convert_xls_to_xlsx(xls_path):
    print('Converting the file...')
    print(xls_path)
    # Open the .xls file
    with open_workbook(xls_path) as xls_book:
        # Create a new .xlsx workbook
        xlsx_book = Workbook()

        # Iterate through all sheets
        for i in range(xls_book.nsheets):
            # Get the worksheet
            xls_sheet = xls_book.sheet_by_index(i)
            
            # Create a new worksheet in xlsx file with the same name
            if i == 0:
                xlsx_sheet = xlsx_book.active
                xlsx_sheet.title = xls_sheet.name
            else:
                xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

            # Copy the cell values and styles
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    xlsx_sheet.cell(row=row+1, column=col+1, value=xls_sheet.cell_value(row, col))

        # Generate the .xlsx filename
        xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'

        # Save the .xlsx file
        xlsx_book.save(xlsx_path)

    return xlsx_path