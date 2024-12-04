import numpy as np
import pandas as pd
import openpyxl as xl
from utilities import convert_xls_to_xlsx
import warnings

warnings.filterwarnings('ignore')

for city_name, excel_file in zip(['blr', 'che', 'del', 'gur', 'mum'], 
                           [r"C:\Users\purus\Downloads\Bliss-blr-oct24(1).xlsx",
                            r"C:\Users\purus\Downloads\onst center Chennai -oct-24.xlsx",
                            r"C:\Users\purus\Downloads\Cost center Delhi-oct-24.xlsx",
                            r"C:\Users\purus\Downloads\Cost center Gurgaon-oct-24.xlsx",
                            r"C:\Users\purus\Downloads\Bliss-mum-oct24.xlsx"]):
    print("Loading the workbooks...")
    try:
        workbook = xl.load_workbook(excel_file)
    except:
        workbook = xl.load_workbook(convert_xls_to_xlsx(excel_file))

    tally = workbook[workbook.sheetnames[0]]

    mis_master = pd.read_excel(r"S:\smoor\data\Master MIS.xlsx")
    mis_master.drop_duplicates(subset=['Code'], inplace=True)

    # Loading the categories
    cc_category = pd.read_excel(r"S:\smoor\data\Master MIS.xlsx", "Mapping",
                                skiprows=[0]).T
    cc_category.reset_index(inplace=True)
    cc_category.columns = cc_category.iloc[0]
    cc_category.drop(0, inplace=True)
    cc_category[cc_category.columns[0]] = cc_category[cc_category.columns[0]].apply(lambda x: np.nan if x.startswith('Unnamed:') else x)

    # Renaming the columns
    print("Renaming the headers...")
    mis_names = pd.read_excel(f"S:/smoor/data/names_{city_name}.xlsx")
    mis_names['Name'] = mis_names['Name'].apply(lambda x: x.strip())
    headers = [c.value for c in next(tally.iter_rows(min_row=3, max_row=3, min_col=13))]
    new_headers = []
    for header in headers:
        try:
            rename = mis_names[mis_names['Name'] == header.strip()]['Rename']
            new_headers.append(rename.values[0])
        except:
            print(header)
            new_headers.append(header)

    # Getting the GL codes and names
    print("Extracting the GL codes...")
    gl_codes, names = [], []
    for col in tally.iter_cols(min_row=4, min_col=5, max_col=5, values_only=True):
        for val in col:
            try:
                if '|' in val:
                    gl_code, name = list(map(lambda x: x.strip(), val.split('|')[:2]))
                elif '-' in val:
                        gl_code, name = list(map(lambda x: x.strip(), val.split('-')[:2]))
                elif ' I ' in val:
                    gl_code, name = list(map(lambda x: x.strip(), val.split(' I ')[:2]))
                gl_codes.append(int(gl_code))
                names.append(name)
            except:
                gl_codes.append(val)
                names.append(val)

    # Getting the Party A/c
    print("Extracting the Party A/c...")
    for col in tally.iter_cols(min_row=4, min_col=6, max_col=6, values_only=True):
        party_ac = col

    # Getting the vch and led
    print("Extracting the VCH and LED Narration...")
    vch_narration, led_narration = [], []
    for col_idx, col in enumerate(tally.iter_cols(min_row=4, min_col=11, max_col=12, values_only=True), start=11):
        if col_idx == 11:
            vch_narration.extend(col)
        elif col_idx == 12:
            led_narration.extend(col)

    # Getting columns F-L
    print("Extrating the master columns...")
    f, g, h, i, j, k, l = [], [], [], [], [], [], []
    lists = [f, g, h, i, j, k, l]
    for code in gl_codes:
        try:
            vals = mis_master[mis_master['Code'] == code].values[0][4:-1]
            for val, list_ in zip(vals, lists):
                list_.append(val)
        except:
            for list_ in lists:
                list_.append(None)

    # Getting the cell values
    print("Populating the rest of the sheet...")
    temp = []
    for idx, col in enumerate(tally.iter_cols(min_row=4, min_col=13)):
        # Getting the column name
        col_name = new_headers[idx]
        for cell in col:
            if cell.value == None:
                temp.append(0)
            elif cell.number_format[-3:-1] == "Dr":
                temp.append(cell.value)
            elif cell.number_format[-3:-1] == "Cr":
                temp.append(-1*float(cell.value))
        if idx == 0:
            value_df = pd.DataFrame({col_name: temp})
        else:
            if col_name in value_df.columns:
                value_df[col_name] += temp
            else:
                value_df[col_name] = temp
        temp = []

    # Separate file needs to be generated!

    # Creating the total column
    value_df['Total'] = value_df.sum(axis=1)

    # Creating the category-wise total
    for col in cc_category.columns:
        cost_centers = list(cc_category[col].dropna().values)
        final_cc = []
        for cc in cost_centers:
            if cc in value_df.columns:
                final_cc.append(cc)
            
        value_df[col] = value_df[final_cc].sum(axis=1)

    master_cols = pd.DataFrame(np.array([gl_codes, names, party_ac, vch_narration, led_narration, f, g, h, i, j, k, l]).T,
                            columns=['GL Code', 'Name', 'Party A/c', 'Vch Narration', 'Led Narration', 'A/c Grp 5', 'A/c Grp 4', 'A/c Grp3 (Alloc)', 'A/c Grp 2 (MIS)', 'A/c Grp 1', 'Verical Grp', 'Final Grp in P&L'])
    final_df = pd.concat([master_cols, value_df],
                        axis=1)
    final_df.drop(final_df.tail(1).index,inplace=True)
    final_df.to_csv(f"{city_name.upper()}_Oct.csv", index=False)
    print(f"Files saved for {city_name}!")